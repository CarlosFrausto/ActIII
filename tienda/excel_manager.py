from openpyxl import Workbook, load_workbook # Se debe instalar con PIP openpyxl...
import os

ARCHIVO = "Registro_productos.xlsx"

ENCABEZADOS = [
    "id",
    "nombre",
    "categoria",
    "precio",
    "stock",
    "ventas_totales"
]


def crear_excel():
    if not os.path.exists(ARCHIVO):
        wb = Workbook()
        ws = wb.active
        ws.append(ENCABEZADOS)
        wb.save(ARCHIVO)


def cargar_productos():
    crear_excel()

    wb = load_workbook(ARCHIVO)
    ws = wb.active

    productos = []

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0] is not None:
            productos.append({
                "id": str(fila[0]),
                "nombre": fila[1],
                "categoria": fila[2],
                "precio": float(fila[3]),
                "stock": int(fila[4]),
                "ventas_totales": int(fila[5])
            })

    return productos


def guardar_productos(lista):
    wb = Workbook()
    ws = wb.active

    ws.append(ENCABEZADOS)

    for p in lista:
        ws.append([
            p["id"],
            p["nombre"],
            p["categoria"],
            p["precio"],
            p["stock"],
            p["ventas_totales"]
        ])

    wb.save(ARCHIVO)